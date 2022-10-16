#!/usr/bin/python3
dict={}
import openpyxl
#함수
def calGrade(list):
    #print(list)
    grade=[]
    grade2=[]
    #sum구하기
    sum=0
    for t in list:
        sum = sum + t[1]
    #학점계산 + 겸사겸사 sumA,sumB,sumC도 구하기
    sumA=0
    sumB=0
    sumC=0
    for i in range (len(list)):
        #per구하기
        sum2=0
        for j in range(i+1):
            sum2 = sum2+list[j][1]
        per=(sum2/sum)*100
        #대략적인 학점주기(A,B,C까지만)
        if per<=30:
            sumA = sumA + list[j][1]
            grade.append("A")
        elif per<=70:
            sumB = sumB + list[j][1]
            grade.append("B")
        else:
            sumC = sumC + list[j][1]
            grade.append("C")
     #학점주기(+,0까지)
    sum2A=0
    sum2B=0
    sum2C=0
    for i in range(len(list)):
        if (grade[i] == 'A'):
            sum2A = sum2A + list[i][1]
            per=(sum2A/sumA)*100
            if per<= 50:
                grade2.append("A+")
            else:
                grade2.append("A0")
        elif (grade[i] == 'B'):
            sum2B = sum2B + list[i][1]
            per=(sum2B/sumB)*100
            if per<= 50:
                grade2.append("B+")
            else:
                grade2.append("B0")
        elif (grade[i] == 'C'):
            sum2C = sum2C + list[i][1]
            per=(sum2C/sumC)*100
            if per<= 50:
                grade2.append("C+")
            else:
                grade2.append("C0")

    return grade2
#엑셀열기
wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1;
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		if sum_v not in dict:
			dict[sum_v] = 1
		else:
			dict[sum_v] += 1
	row_id += 1
#print(dict)
# grade계산
scoreList = list(dict.items())
scoreList.sort(key=lambda x:-x[0])
grade=calGrade(scoreList)
#print(grade)
#성적_학점 딕셔너리 만들기
score_grade_dict={}
for i in range(len(scoreList)):
    if scoreList[i][0] not in score_grade_dict:
        score_grade_dict[scoreList[i][0]] = grade[i]
#print(score_grade_dict)
# 채워넣기(성적입력>학점출력)
row_id = 1
for row in ws:
	if row_id != 1:
		ws.cell(row = row_id, column = 8).value = score_grade_dict[ws.cell(row = row_id, column = 7).value]
	row_id += 1
#저장하기
wb.save("student.xlsx")

